import itertools
import math
import re

from openpyxl import Workbook, cell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN

from run_quine_mccluskey import BFV
from texttable import Texttable


class Cell:
    def __init__(self, s: str, co: bool):
        self.s = s
        self.crossed_out = co
        self.glued_up = False

    def __repr__(self) -> str:
        return self.s

    def __len__(self) -> int:
        return len(self.s)

    def __eq__(self, other):
        return self.s == other.s

    def k_str(self, ids: str) -> str:
        return f'K({"".join([("!" if self.s[i] == "0" else "") + ind for i, ind in enumerate(ids)])})'

    @staticmethod
    def lhs_can_be_overwritten_by_rhs(lhs_s: str, lhs_ids: str, rhs_s: str, rhs_ids: str) -> bool:
        if not len(rhs_s) == len(lhs_s) - 1:
            return False

        more = 0
        for i, id in enumerate(lhs_ids):
            if id not in rhs_ids:
                more += 1
                continue

            if not lhs_s[i] == rhs_s[rhs_ids.find(id)]:
                return False

        if not more == 1:
            return False
        return True


def glue_table(tab: list[list[Cell]]) -> None:
    for i in range(1, len(tab)):
        for j in range(len(tab[i]) - 2, -1, -1):
            if tab[i][j].crossed_out or tab[i][j].glued_up:
                continue
            for check_j in range(len(tab[i]) - 2, -1, -1):
                if tab[i][check_j].crossed_out:
                    continue
                if Cell.lhs_can_be_overwritten_by_rhs(
                        tab[i][j].s, tab[0][j].s,
                        tab[i][check_j].s, tab[0][check_j].s):
                    tab[i][j].glued_up = True
                    break


def get_table(tab: list[list[Cell]]) -> str:
    t = Texttable()
    t.set_chars(['—', '|', '+', '—'])
    t.set_cols_dtype(['t'] * len(tab[0]))
    t.set_cols_align(['c'] * len(tab[0]))
    t.set_cols_width([len(c) for c in tab[0]])
    t.set_deco(Texttable.BORDER | Texttable.HEADER | Texttable.VLINES)
    t.add_rows(tab)
    return t.draw()


def get_equations(tab: list[list[Cell]]) -> str:
    def line_has_1s(l: list[Cell]) -> bool:
        for j in range(len(l) - 1):
            if not l[j].crossed_out:
                return True
        return False

    out_s = str()
    for i in range(1, len(tab)):
        if not line_has_1s(tab[i]):
            continue

        first = True
        for j in range(len(tab[i]) - 1):
            if tab[i][j].crossed_out or tab[i][j].glued_up:
                continue

            out_s += ' v' if not first else ''
            first = False
            out_s += tab[i][j].k_str(tab[0][j].s)

        out_s += ' = 1\n'

    return out_s


def main():
    table = list[list[Cell]]()
    n = int(math.log2(len(BFV)))
    assert n > 0

    header = list[Cell]()
    for k in range(1, n + 1):
        combs = itertools.combinations(range(n), k)
        header.extend([Cell(''.join(list(map(str, comb))), False) for comb in combs])

    table.append(header + [Cell('F', False)])

    sets = list(map(lambda s: list(map(lambda x: Cell(str(x), False), s)),
                    list(itertools.product([0, 1], repeat=n))))

    table.extend(sets)

    for j in range(n, len(table[0]) - 1):
        indexes = re.split('', table[0][j].s)
        indexes.pop(0)
        indexes.pop(-1)
        indexes = list(map(int, indexes))

        for i in range(1, len(table)):
            cur_set = list(map(lambda ind: str(table[i][ind]), indexes))
            table[i].append(Cell(''.join(cur_set), False))

    for i in range(1, len(table)):
        table[i].append(Cell(str(BFV[i - 1]), False))

    # Crossing out lines F = 0
    for i in range(1, len(table)):
        if table[i][-1].s == '1':
            continue
        for j in range(len(table[i])):
            table[i][j].crossed_out = True

    # Crossing out excess in columns
    for j in range(len(table[0])):
        for i in range(1, len(table)):
            if not table[i][j].crossed_out:
                continue

            for secondary_i in range(1, len(table)):
                if table[secondary_i][j].s == table[i][j].s:
                    table[secondary_i][j].crossed_out = True

    with open('mc-table.txt', 'w') as writer:
        writer.write(get_table(table))

    wb = Workbook()
    ws = wb.active

    for row in table:
        ws.append(list(map(str, row)))

    font = Font(bold=False, color='000000')
    al = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(border_style=BORDER_THIN,
                              color='000000'),
                    right=Side(border_style=BORDER_THIN,
                               color='000000'))
    ok_fill = PatternFill('solid', fgColor='00FFFFFF')

    bad_fill = PatternFill('solid', fgColor='00C0C0C0')

    ag_fill = PatternFill('solid', fgColor='009999FF')

    for j in range(len(table[0])):
        c = cell.cell.Cell(ws, row=1, column=j + 1)
        ws.column_dimensions[c.column_letter].width = len(table[0][j]) + 1

    for i in range(1, len(table)):
        for j in range(len(table[i])):
            c = cell.cell.Cell(ws, row=i + 1, column=j + 1)
            ws[c.coordinate].alignment = al
            ws[c.coordinate].border = border

            if table[i][j].crossed_out:
                ws[c.coordinate].fill = bad_fill
                ws[c.coordinate].font = font
                continue

            ws[c.coordinate].fill = ok_fill
            ws[c.coordinate].font = font

    wb.save('mc-table-crossed.xlsx')

    glue_table(table)

    for i in range(1, len(table)):
        for j in range(len(table[i]) - 1):
            c = cell.cell.Cell(ws, row=i + 1, column=j + 1)

            if not table[i][j].crossed_out and not table[i][j].glued_up:
                ws[c.coordinate].fill = ag_fill

    wb.save('mc-table-crossed-glued.xlsx')

    with open('mc-eqs.txt', 'w') as writer:
        writer.write(get_equations(table))


if __name__ == '__main__':
    main()
